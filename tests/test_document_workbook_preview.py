"""Read-only Excel workbook preview for the admin document-review workflow.

Covers the openpyxl-backed preview parser (sheets, cells, dates/numbers/text, truncation, safe failure,
source file untouched) and the View-URL routing (Excel -> preview, others -> existing inline route).
"""
from datetime import date

import openpyxl

from app.routes.admin import _view_url
from app.routes.documents import _PREVIEW_MAX_COLS, _PREVIEW_MAX_ROWS, read_workbook_preview


def _make_xlsx(tmp_path, name="expenses.xlsx"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(["Name", "Amount", "Date"])
    ws.append(["Alice", 1234.5, date(2021, 4, 15)])
    ws.append(["Bob", 10, "text"])
    second = wb.create_sheet("Second")
    second.append(["only", "second"])
    f = tmp_path / name
    wb.save(f)
    return f


def test_preview_renders_sheets_and_cell_values(tmp_path):
    r = read_workbook_preview(_make_xlsx(tmp_path))
    assert r["sheetnames"] == ["Sheet1", "Second"]
    assert r["active"] == "Sheet1"
    assert r["rows"][0] == ["Name", "Amount", "Date"]
    assert r["rows"][1][0] == "Alice"
    assert "1234.5" in r["rows"][1][1]          # number
    assert r["rows"][1][2] == "2021-04-15"       # date formatted
    assert r["rows"][2] == ["Bob", "10", "text"]  # int + text


def test_preview_can_switch_worksheet(tmp_path):
    r = read_workbook_preview(_make_xlsx(tmp_path), sheet="Second")
    assert r["active"] == "Second"
    assert r["rows"][0] == ["only", "second"]


def test_preview_truncates_large_workbook(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    for _ in range(_PREVIEW_MAX_ROWS + 50):
        ws.append(list(range(_PREVIEW_MAX_COLS + 10)))
    f = tmp_path / "big.xlsx"
    wb.save(f)
    r = read_workbook_preview(f)
    assert r["truncated_rows"] is True and r["truncated_cols"] is True
    assert len(r["rows"]) == _PREVIEW_MAX_ROWS
    assert len(r["rows"][0]) == _PREVIEW_MAX_COLS


def test_preview_does_not_modify_source_file(tmp_path):
    f = _make_xlsx(tmp_path)
    before = f.read_bytes()
    read_workbook_preview(f)
    read_workbook_preview(f, sheet="Second")
    assert f.read_bytes() == before              # strictly read-only


def test_preview_fails_safely_on_bad_workbook(tmp_path):
    f = tmp_path / "bad.xlsx"
    f.write_bytes(b"this is not a real workbook")
    r = read_workbook_preview(f)
    assert "error" in r                          # safe failure, no crash


def test_view_url_routes_excel_to_preview_others_to_inline():
    assert _view_url(5, "Expenses.xlsx") == "/documents/5/preview"
    assert _view_url(5, "book.xlsm") == "/documents/5/preview"
    assert _view_url(7, "2021 8879 S.pdf") == "/documents/7/download?inline=1"   # PDF unchanged
    assert _view_url(7, "IMG_5178.HEIC") == "/documents/7/image-preview"          # HEIC -> image preview
    assert _view_url(7, "notes.docx") == "/documents/7/download?inline=1"        # other -> existing behavior


def test_xlsx_download_remains_attachment_not_inline():
    from app.routes.documents import _is_inline_viewable
    # The Download button hits /documents/{id}/download (no inline) -> attachment; even ?inline=1 refuses
    # to render Excel inline, so Download always returns the original file.
    assert _is_inline_viewable("application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               "x.xlsx") is False
