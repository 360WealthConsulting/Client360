"""Read-only helpers for bounded Excel workbook previews."""

from datetime import date, datetime, time


def _fmt_cell(value):
    """Render a workbook cell value for read-only display."""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == time(0, 0) else value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def read_workbook_preview(path, sheet="", *, max_rows, max_cols):
    """Read a bounded, read-only preview of an .xlsx/.xlsm workbook."""
    import openpyxl

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001
        return {"error": f"This workbook could not be opened for preview ({type(exc).__name__})."}

    try:
        sheetnames = list(wb.sheetnames)
        active = sheet if sheet in sheetnames else (sheetnames[0] if sheetnames else None)
        rows, truncated_rows, truncated_cols = [], False, False

        if active is not None:
            ws = wb[active]
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx >= max_rows:
                    truncated_rows = True
                    break

                cells = []
                for c_idx, val in enumerate(row):
                    if c_idx >= max_cols:
                        truncated_cols = True
                        break
                    cells.append(_fmt_cell(val))

                rows.append(cells)

        return {
            "sheetnames": sheetnames,
            "active": active,
            "rows": rows,
            "truncated_rows": truncated_rows,
            "truncated_cols": truncated_cols,
        }
    finally:
        wb.close()
