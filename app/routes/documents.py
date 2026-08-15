from pathlib import Path

from fastapi import APIRouter, File, Form, Request, UploadFile
from fastapi.responses import FileResponse, HTMLResponse, RedirectResponse
from fastapi.templating import Jinja2Templates
from sqlalchemy import select

from app.db import engine, people
from app.services.documents import (
    archive_document,
    get_document,
    get_person_documents,
    save_person_document,
)
from app.services.microsoft_documents import get_person_microsoft_documents
from app.services.timeline import add_timeline_event
from app.templating import render_error

router = APIRouter()
templates = Jinja2Templates(directory="app/templates")


@router.get(
    "/people/{person_id}/documents",
    response_class=HTMLResponse,
)
def person_documents(request: Request, person_id: int):
    with engine.connect() as connection:
        person = connection.execute(
            select(people).where(people.c.id == person_id)
        ).mappings().one_or_none()

    if person is None:
        return HTMLResponse(
            "<h1>Person not found</h1>",
            status_code=404,
        )

    return templates.TemplateResponse(
        request=request,
        name="people/documents.html",
        context={
            "person": person,
            "documents": get_person_documents(person_id),
            "microsoft_documents": get_person_microsoft_documents(person_id),
            "uploaded": request.query_params.get("uploaded") == "1",
            "archived": request.query_params.get("archived") == "1",
        },
    )


@router.post("/people/{person_id}/documents")
async def upload_person_document(
    person_id: int,
    file: UploadFile = File(...),
    category: str = Form("other"),
    description: str = Form(""),
    uploaded_by: str = Form(""),
):
    with engine.connect() as connection:
        person_exists = connection.execute(
            select(people.c.id).where(people.c.id == person_id)
        ).scalar_one_or_none()

    if person_exists is None:
        return HTMLResponse(
            "<h1>Person not found</h1>",
            status_code=404,
        )

    if not file.filename:
        return HTMLResponse(
            "<h1>A file is required</h1>",
            status_code=400,
        )

    document_id = save_person_document(
        person_id=person_id,
        original_name=file.filename,
        source=file.file,
        content_type=file.content_type,
        category=category,
        description=description,
        uploaded_by=uploaded_by,
    )

    await file.close()

    add_timeline_event(
        person_id=person_id,
        source="client360",
        event_type="document_uploaded",
        title="Document Uploaded",
        summary=file.filename,
        external_id=f"document-uploaded-{document_id}",
        event_metadata={
            "document_id": document_id,
            "category": category,
            "description": description or None,
            "uploaded_by": uploaded_by or None,
            "content_type": file.content_type,
        },
    )

    return RedirectResponse(
        url=f"/people/{person_id}/documents?uploaded=1",
        status_code=303,
    )


def _is_inline_viewable(content_type: str | None, name: str | None) -> bool:
    """Types safe to render inline in the browser (PDF / image / plain text). Everything else
    downloads. Falls back to the filename extension when no content_type is recorded."""
    ct = (content_type or "").lower()
    if ct == "application/pdf" or ct.startswith("image/") or ct == "text/plain":
        return True
    ext = (name or "").rsplit(".", 1)[-1].lower() if "." in (name or "") else ""
    return ext in {"pdf", "png", "jpg", "jpeg", "gif", "webp", "tif", "tiff", "bmp", "txt",
                   "heic", "heif"}


@router.get("/documents/{document_id}/download")
def download_document(document_id: int, request: Request, inline: bool = False):
    document = get_document(document_id)

    if document is None or document["archived"]:
        return render_error(request, 404,
                            detail="This document is no longer available. It may have been archived.")

    # Serve the absolute storage_uri whenever the document carries one — this covers every durable
    # canonical copy (TaxDome-synced "Client360 Local" AND relocated "Client360 Repository" documents,
    # whose storage_path is stored relative to the content root). Directly-uploaded legacy documents have
    # no absolute storage_uri and continue to resolve via their repo-relative storage_path.
    if document["storage_uri"] and Path(document["storage_uri"]).is_absolute():
        path = Path(document["storage_uri"])
    else:
        path = Path(document["storage_path"])

    if not path.exists():
        return render_error(request, 404,
                            detail="The stored copy of this document could not be found on the server.")

    # ?inline=1 renders viewable types (PDF/image/text) in the browser so an operator can inspect a
    # document without downloading it. Authorization is unchanged (enforced by the middleware on this
    # path). Non-viewable types always download.
    disposition = "inline" if (inline and _is_inline_viewable(
        document["content_type"], document["original_name"])) else "attachment"
    return FileResponse(
        path=path,
        media_type=document["content_type"] or "application/octet-stream",
        filename=document["original_name"],
        content_disposition_type=disposition,
    )


_EXCEL_EXTS = {"xlsx", "xlsm"}          # openpyxl reads these; legacy .xls is not supported here
_PREVIEW_MAX_ROWS = 200
_PREVIEW_MAX_COLS = 30
_PREVIEW_MAX_FILE_BYTES = 25 * 1024 * 1024


def _fmt_cell(value):
    """Render a workbook cell value for read-only display (dates/numbers/text handled reasonably)."""
    from datetime import date, datetime, time
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat() if value.time() == time(0, 0) else value.isoformat(sep=" ")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def read_workbook_preview(path, sheet=""):
    """Read a bounded, READ-ONLY preview of an .xlsx/.xlsm workbook. Never writes or converts the file.
    Returns {sheetnames, active, rows, truncated_rows, truncated_cols} or {error} on failure."""
    import openpyxl
    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:  # noqa: BLE001 — malformed/unreadable workbook: fail safely
        return {"error": f"This workbook could not be opened for preview ({type(exc).__name__})."}
    try:
        sheetnames = list(wb.sheetnames)
        active = sheet if sheet in sheetnames else (sheetnames[0] if sheetnames else None)
        rows, truncated_rows, truncated_cols = [], False, False
        if active is not None:
            ws = wb[active]
            for r_idx, row in enumerate(ws.iter_rows(values_only=True)):
                if r_idx >= _PREVIEW_MAX_ROWS:
                    truncated_rows = True
                    break
                cells = []
                for c_idx, val in enumerate(row):
                    if c_idx >= _PREVIEW_MAX_COLS:
                        truncated_cols = True
                        break
                    cells.append(_fmt_cell(val))
                rows.append(cells)
        return {"sheetnames": sheetnames, "active": active, "rows": rows,
                "truncated_rows": truncated_rows, "truncated_cols": truncated_cols}
    finally:
        wb.close()


@router.get("/documents/{document_id}/preview")
def preview_document(document_id: int, request: Request, sheet: str = ""):
    """Read-only Client360 workbook preview for .xlsx/.xlsm documents, so an admin can inspect a
    spreadsheet in a new tab instead of downloading it. Authorization is unchanged — the middleware
    enforces the same document-scope rules on this ``/documents/{id}`` path (including the admin
    unassigned-document exception). Never modifies the source file, metadata, or ownership."""
    document = get_document(document_id)
    if document is None or document["archived"]:
        return render_error(request, 404,
                            detail="This document is no longer available. It may have been archived.")
    name = document["original_name"] or ""
    ext = name.rsplit(".", 1)[-1].lower() if "." in name else ""
    ctx = {"document_id": document_id, "filename": name,
           "download_url": f"/documents/{document_id}/download"}
    if ext not in _EXCEL_EXTS:
        ctx["error"] = "Preview is available for .xlsx/.xlsm workbooks only. Use Download for this file."
        return templates.TemplateResponse(request=request, name="documents/workbook_preview.html",
                                          context={**ctx, "sheetnames": [], "rows": []})

    if document["storage_uri"] and Path(document["storage_uri"]).is_absolute():
        path = Path(document["storage_uri"])
    else:
        path = Path(document["storage_path"])
    if not path.exists():
        return render_error(request, 404,
                            detail="The stored copy of this document could not be found on the server.")
    if path.stat().st_size > _PREVIEW_MAX_FILE_BYTES:
        ctx["error"] = "This workbook is too large to preview safely. Use Download instead."
        return templates.TemplateResponse(request=request, name="documents/workbook_preview.html",
                                          context={**ctx, "sheetnames": [], "rows": []})

    result = read_workbook_preview(path, sheet=sheet)
    return templates.TemplateResponse(request=request, name="documents/workbook_preview.html",
                                      context={**ctx, **result,
                                               "max_rows": _PREVIEW_MAX_ROWS, "max_cols": _PREVIEW_MAX_COLS})


@router.post(
    "/people/{person_id}/documents/{document_id}/archive"
)
def archive_person_document(
    person_id: int,
    document_id: int,
):
    if not archive_document(document_id, person_id):
        return HTMLResponse(
            "<h1>Document not found</h1>",
            status_code=404,
        )

    return RedirectResponse(
        url=f"/people/{person_id}/documents?archived=1",
        status_code=303,
    )
